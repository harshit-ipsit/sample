#!/usr/bin/python

import streamlit as st
import threading
from openpyxl import load_workbook
import pandas as pd


def threaded(fn):
    def wrapper(*args, **kwargs):
        thread = threading.Thread(target=fn, args=args, kwargs=kwargs)
        thread.start()
        return thread

    return wrapper


# --- Page Config ---
st.set_page_config(
    page_title="KTR Updates",
    page_icon=":orange_book:",
    layout="wide",
    initial_sidebar_state="expanded",
)

# --- Inline Custom CSS ---
st.markdown("""
    <style>
    /* Hide Streamlit Branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    [data-testid="stSidebarNav"] {
        background-image: url(https://w.amazon.com/rest/wikis/xwiki/s3files/file/logo-generic.png);
        background-repeat: no-repeat;
        padding-top: 10px;
        background-position: 20px 20px;
    }
    </style>
""", unsafe_allow_html=True)

# --- List Declaration ---
# Summary
project_poc_list = []
project_name_list = []
project_description_list = []
current_testing_phase_list = []
current_testing_phase_eta_list = []

# Automation
test_case_feasibility_count_list = []
test_case_scripting_count_list = []
test_scripts_cr_count_list = []
pipeline_integration_type_list = []
test_case_feasibility_status_list = []
test_case_scripting_status_list = []
test_scripts_cr_status_list = []
pipeline_integration_status_list = []
test_case_feasibility_eta_list = []
test_case_scripting_eta_list = []
test_scripts_cr_eta_list = []
pipeline_integration_eta_list = []

# Additional Information
slow_spots_list = []
low_lights_list = []
one_or_two_way_decision_list = []
important_callouts_list = []

# Issues Information
issues_high_count_list = []
launch_blockers_details_list = []
issues_total_count_list = []
issues_medium_count_list = []
issues_low_count_list = []

# Test Planning Information
requirement_analysis_status_list = []
requirement_analysis_eta_list = []
use_case_created_count_list = []
test_case_created_count_list = []
use_case_created_status_list = []
test_case_created_status_list = []
use_case_created_eta_list = []
test_case_created_eta_list = []

# General Information
test_data_generation_count_list = []
adhoc_to_tc_conversion_count_list = []
adhoc_testing_bugs_found_count_list = []
bugs_verified_count_list = []
test_case_optimized_count_list = []
test_data_generation_status_list = []
adhoc_to_tc_conversion_status_list = []
adhoc_testing_bugs_found_status_list = []
bugs_verified_status_list = []
test_case_optimized_status_list = []
test_data_generation_eta_list = []
adhoc_to_tc_conversion_eta_list = []
adhoc_testing_bugs_found_eta_list = []
bugs_verified_eta_list = []
test_case_optimized_eta_list = []

# st.title("KTR Updates: WK-48 & WK-49")
original_title = '<p style="font-family:Calibri; color:#FF6600; font-size: 40px;">' \
                 'KTR Updates: WK-48 & WK-49</p>'
st.markdown(original_title, unsafe_allow_html=True)
st.markdown("""---""")

with st.container():
    # Project POC:
    project_poc = st.selectbox(
        'Project POC',
        ('Ramanagiri', 'Aravindan', 'Arun', 'Sandhiya', 'Ashwin', 'Janani', 'Bharath', 'Gowtham', 'Manothini', 'Nancy',
         'Paranthaman', 'Nelson', 'Keerthika', 'Suresh Krishna', 'Sudharsanan', 'Kubra', 'Saranya', 'Shubham',
         'Shrijith', 'Siva', 'Sunaina', 'Swetha', 'Varadha', 'Subashree'))

with st.container():
    if project_poc == "Ramanagiri":
        project = st.selectbox(
            'Project Name',
            ('TSW Missing Strings Audit', 'Release Day', 'KEP Self Service Proofs/Samples Releases',
             'KEP POD for Onix Beta Extension', 'KEP POD Royalties', 'Automation: Visual Testing for Release Day',
             'Automation: TSW Missing Strings Audit', 'Automation: TSW Failure Fixes for NA',
             'Automation: JP Failure Fixes for Bookshelf and TSW'))
    if project_poc == "Aravindan":
        project = st.selectbox('Project Name', ('DER Seller Automation Failure Fixes', 'NA'))
    if project_poc == "Arun":
        project = st.selectbox('Project Name', (
            'Download Button for Authors Report [Estimate code finish time: Dec 12th]', 'Author CX',
            'Episode Discussion Refactoring Regression Pass', 'Token / Acquisition Offer Promo Offers'))
    if project_poc == "Sandhiya":
        project = st.selectbox('Project Name', ('KDP Marketing Website', 'Automation: JP Manga',
                                                'Automation: Automated Authorization Checks for KDP Marketing Pages'))
    if project_poc == "Ashwin":
        project = st.selectbox('Project Name', (
            'CSP Migration Website Horizonte 5.3 Upgrade Testing', 'OD3/PD2 - 7/31 KDP BI', 'KDP Categories',
            'HAEM OD3/PD2 QA Engagement'))
    if project_poc == "Janani":
        project = st.selectbox('Project Name', (
            'Engagement Offer Testing', 'Promo Productization Testing',
            'Falkor Reader Pre-Prod Automation - QA Pipeline'))
    if project_poc == "Bharath":
        project = st.selectbox('Project Name',
                               ('OD3/PD2 KDP CS BI', 'CSP OD3 Testing Campaign', 'POD BI Privacy QA Testing'))
    if project_poc == "Gowtham":
        project = st.selectbox('Project Name', (
            'Ignite Regression Tasks', 'Request to build course player mobile web browser application test suite'))
    if project_poc == "Manothini":
        project = st.selectbox('Project Name', ('KDP Banner Tool Multi-Tenant', 'NA'))
    if project_poc == "Nancy":
        project = st.selectbox('Project Name',
                               ('POD Manufacturing Fee Changes', 'Automation: New Print Publishing Languages'))
    if project_poc == "Paranthaman":
        project = st.selectbox('Project Name', ('Test Case Optimization - Self Service Series / Potter '
                                                '(E-Book and Paperback) / Hardcover', 'NA'))
    if project_poc == "Nelson":
        project = st.selectbox('Project Name', (
            'KDP Customer Feedback Widget', 'Hardcover ATOM', 'Transparency Phase 2', 'MAS Regression Dec',
            'MAS Regression Nov', 'Quick Sight Dasboard Implementation for Testrail', 'MCM Audit Process',
            'Automation: Visual Testing - Create Space Migration', 'Automation: Visual Testing - Title Archiving',
            'Automation: Visual Testing - Release Day'))
    if project_poc == "Keerthika":
        project = st.selectbox('Project Name', (
            'Promotion Price Lock Validation for Digital Books', 'Title Archiving for KDP Bookshelf',
            'IPX Privacy ASIN Takedowns (Non-Deletion)'))
    if project_poc == "Suresh Krishna":
        project = st.selectbox('Project Name', (
            'CDS PD2', 'Privacy Testing for KDP Drone Integration (Away Team Support)',
            'Upgrade Amazon Connect in Book Support Salesforce Instance', 'Goodreads Community Bridge', 'CRS2',
            'Multi-Tenant CDS', 'Forge OD3/PD2 Test Effort', 'Jenkins Implementation'))
    if project_poc == "Sudharsanan":
        project = st.selectbox('Project Name', ('Tandem (Confidential Project)', 'C4A Migration Testing',
                                                'Automation: Extending JP Marketplace support for the API'
                                                ' test framework',
                                                'Automation: Hardcover Regression Suite Failure Fixes'))
    if project_poc == "Kubra":
        project = st.selectbox('Project Name', (
            'KDP Indie Manga Pricing 12/30', 'Automation: Series QA Pipeline Failure Fixes',
            'Automation: Title Archiving for KDP Bookshelf Automation'))
    if project_poc == "Saranya":
        project = st.selectbox('Project Name', ('Bulk Unlock CX', 'NA'))
    if project_poc == "Shubham":
        project = st.selectbox('Project Name', (
            'Indie Tateyomi for KDP JP', 'Automation: Self Service Series Automated Test Scripts Updation',
            'Automation: Author Copies Failure Fixes'))
    if project_poc == "Shrijith":
        project = st.selectbox('Project Name', ('AOS - The Broken Doll Launch and Surface Testing',
                                                'AOS - Regression Tests for Bulk Borrow in DBS Master Suite',
                                                'AOS Launch and Surface Testing', 'SIM SNS',
                                                'Automation: AB4B - QA Pipeline Failure Fixes',
                                                'Automation: Series NA Automation'))
    if project_poc == "Siva":
        project = st.selectbox('Project Name', ('TSW React Print ISBN Component',
                                                'PD2 Campaign - FAKT', 'New Print Royalty Calculator'))
    if project_poc == "Sunaina":
        project = st.selectbox('Project Name', ('KDP Series Kraken API', 'Automation: KDP Print Expansion'))
    if project_poc == "Swetha":
        project = st.selectbox('Project Name', ('Support for Comixology', 'NA'))
    if project_poc == "Varadha":
        project = st.selectbox('Project Name', (
            'Replace Order Proof Copies Banner with Card', 'Create Space Migration Workflow Changes'))
    if project_poc == "Subashree":
        project = st.selectbox('Project Name', (
            'Book Support - Sensitive Data Detection',
            'Book Support - Auto Assign CRAs, reopen, and put into WIP',
            'Book Support - Automate Phone calls for Book Support Agents'))

with st.container():
    project_description = st.text_area("Project Description ("
                                       "* Specify the Region, Format, Etc. for Automation related Updates or "
                                       "Short Description of New Projects!)")

with st.container():
    col1, col2 = st.columns([1, 3])
    with col1:
        current_testing_phase = st.selectbox('Current Testing Phase', (
            'Functional Testing / 1st Round', 'Bug Fixes', 'Code Drop', 'Bug Fix Verification',
            'Regression / 2nd Round of Testing', 'Final Round of Testing', 'Project Launch', 'Automation',
            'Prod Sanity', 'QA Sign Off'), key="25")
    with col2:
        current_testing_phase_eta = st.text_input(
            'ETA / Comments / Launch Date (* to be provided if phase is In-Progress or '
            'Waiting / Blocked / Launched)',
            key="26")

if "Automation" in project:
    automation_information = st.tabs(["Automation Information"])
    with st.container():
        col1, col2, col3 = st.columns(3)
        with col1:
            test_case_feasibility_count = st.text_input('Test Case Feasibility Check for Automation Count')
            test_case_scripting_count = st.text_input('Test Case Scripting Count')
            test_scripts_cr_count = st.text_input('Test Scripts Count Submitted for Code Review')
            pipeline_integration_type = st.selectbox('Pipeline Integration Type', ('NA', 'DEV', 'QA'))
        with col2:
            test_case_feasibility_status = st.selectbox('Status', ('NA', 'In-Progress', 'On-Hold / Blocked',
                                                                   'Completed'), key="1")
            test_case_scripting_status = st.selectbox('Status', ('NA', 'In-Progress', 'On-Hold / Blocked',
                                                                 'Completed'), key="2")
            test_scripts_cr_status = st.selectbox('Status', ('NA', 'In-Progress', 'On-Hold / Blocked',
                                                             'Completed'), key="3")
            pipeline_integration_status = st.selectbox('Status', ('NA', 'In-Progress',
                                                                  'On-Hold / Blocked', 'Completed'), key="4")
        with col3:
            test_case_feasibility_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="5")
            test_case_scripting_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="6")
            test_scripts_cr_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="7")
            pipeline_integration_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="8")
else:
    test_planning_activity, general_activity, issue_raised_information, additional_information = \
        st.tabs(["Test Planning Activity", "General Activity", "Issue Raised Information", "Additional Information"])
    with additional_information:
        slow_spots = st.text_area('Slow Spots (* if any)')
        low_lights = st.text_area('Low Lights (* if any)')
        one_or_two_way_decision = st.text_area('1 Way / 2 Way Door Decisions (* if any)')
        important_callouts = st.text_area('Important Callouts (* if any)')

    with issue_raised_information:
        col1, col2 = st.columns([3, 1])
        with col1:
            issues_high_count = st.text_input('Issues High Count')
        with col2:
            launch_blockers_details = st.text_input('Launch Blockers (* if any)')
        issues_total_count = st.text_input('Issues Total Count')
        issues_medium_count = st.text_input('Issues Medium Count')
        issues_low_count = st.text_input('Issues Low Count')

    with general_activity:
        col1, col2, col3 = st.columns(3)
        with col1:
            test_data_generation_count = st.text_input('Test Data Generation Count (* if applicable)')
            adhoc_to_tc_conversion_count = st.text_input('Ad-Hoc to Test Case Converted Count (* if applicable)')
            adhoc_testing_bugs_found_count = st.text_input('Ad-Hoc Testing - Bugs Found Count (* if applicable)')
            bugs_verified_count = st.text_input('Bugs Verified Count (* if applicable)')
            test_case_optimized_count = st.text_input('Test Case Optimized Count (* if applicable)')
        with col2:
            test_data_generation_status = st.selectbox('Status',
                                                       ('NA', 'In-Progress', 'On-Hold / Blocked', 'Completed'), key="9")
            adhoc_to_tc_conversion_status = st.selectbox('Status',
                                                         ('NA', 'In-Progress', 'On-Hold / Blocked', 'Completed'),
                                                         key="10")
            adhoc_testing_bugs_found_status = st.selectbox('Status', ('NA', 'In-Progress', 'On-Hold / Blocked',
                                                                      'Completed'), key="11")
            bugs_verified_status = st.selectbox('Status', ('NA', 'In-Progress', 'On-Hold / Blocked', 'Completed'),
                                                key="12")
            test_case_optimized_status = st.selectbox('Status', ('NA', 'In-Progress', 'On-Hold / Blocked', 'Completed'),
                                                      key="13")
        with col3:
            test_data_generation_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="14")
            adhoc_to_tc_conversion_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="15")
            adhoc_testing_bugs_found_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)',
                key="16")
            bugs_verified_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="17")
            test_case_optimized_eta = st.text_input(
                'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="18")

    with test_planning_activity:
        with st.container():
            col1, col2 = st.columns([3, 1])
            with col1:
                requirement_analysis_status = st.selectbox('Requirement Analysis Status',
                                                           ('NA', 'In-Progress', 'On-Hold / Blocked', 'Completed'),
                                                           key="19")
            with col2:
                requirement_analysis_eta = st.text_input(
                    'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="20")

        with st.container():
            col1, col2, col3 = st.columns(3)
            with col1:
                use_case_created_count = st.text_input('Use Case Created Count')
                test_case_created_count = st.text_input('Test Case Created Count')
            with col2:
                use_case_created_status = st.selectbox('Status',
                                                       ('NA', 'In-Progress', 'On-Hold / Blocked', 'Completed'),
                                                       key="21")
                test_case_created_status = st.selectbox('Status',
                                                        ('NA', 'In-Progress', 'On-Hold / Blocked', 'Completed'),
                                                        key="22")
            with col3:
                use_case_created_eta = st.text_input(
                    'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="23")
                test_case_created_eta = st.text_input(
                    'ETA / Comments (* to be provided if status marked In-Progress / Blocked)', key="24")


@threaded
def submit_data():
    if "Automation" in project:
        # Summary
        project_poc_list.append(project_poc)
        project_name_list.append(project)
        project_description_list.append(project_description)
        current_testing_phase_list.append(current_testing_phase)
        current_testing_phase_eta_list.append(current_testing_phase_eta)

        # Automation
        test_case_feasibility_count_list.append(test_case_feasibility_count)
        test_case_scripting_count_list.append(test_case_scripting_count)
        test_scripts_cr_count_list.append(test_scripts_cr_count)
        pipeline_integration_type_list.append(pipeline_integration_type)
        test_case_feasibility_status_list.append(test_case_feasibility_status)
        test_case_scripting_status_list.append(test_case_scripting_status)
        test_scripts_cr_status_list.append(test_scripts_cr_status)
        pipeline_integration_status_list.append(pipeline_integration_status)
        test_case_feasibility_eta_list.append(test_case_feasibility_eta)
        test_case_scripting_eta_list.append(test_case_scripting_eta)
        test_scripts_cr_eta_list.append(test_scripts_cr_eta)
        pipeline_integration_eta_list.append(pipeline_integration_eta)

        columns_0 = ['Project POC', 'Project Name', 'Project Description', 'Current Testing Phase',
                     'Current Testing Phase ETA', 'Test Case Feasibility Automation Count',
                     'Test Case Feasibility Status', 'Test Case Feasibility ETA', 'Test Case Scripting Count',
                     'Test Case Scripting Status', 'Test Case Scripting ETA', 'Test Scripts Count Submitted for CR',
                     'Test Scripts Submitted for CR Status', 'Test Scripts Submitted for CR ETA',
                     'Pipeline Integration Type', 'Pipeline Integration Status', 'Pipeline Integration ETA']
        df = pd.DataFrame(list(zip(project_poc_list, project_name_list, project_description_list,
                                   current_testing_phase_list, current_testing_phase_eta_list,
                                   test_case_feasibility_count_list, test_case_feasibility_status_list,
                                   test_case_feasibility_eta_list, test_case_scripting_count_list,
                                   test_case_scripting_status_list, test_case_scripting_eta_list,
                                   test_scripts_cr_count_list, test_scripts_cr_status_list,
                                   test_scripts_cr_eta_list, pipeline_integration_type_list,
                                   pipeline_integration_status_list, pipeline_integration_eta_list)), columns=columns_0)
        filename = "C:\\Users\\sureraja\\3D Objects\\sureraja\\ktr_automation_updates.xlsx"
        workbook = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = workbook
        writer.sheets = {ws.title: ws for ws in workbook.worksheets}

        df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)

        writer.close()
        print("Automation Data written successfully to Excel!")
    else:
        # Summary
        project_poc_list.append(project_poc)
        project_name_list.append(project)
        project_description_list.append(project_description)
        current_testing_phase_list.append(current_testing_phase)
        current_testing_phase_eta_list.append(current_testing_phase_eta)

        # Additional Information
        slow_spots_list.append(slow_spots)
        low_lights_list.append(low_lights)
        one_or_two_way_decision_list.append(one_or_two_way_decision)
        important_callouts_list.append(important_callouts)

        # Issues Information
        issues_high_count_list.append(issues_high_count)
        launch_blockers_details_list.append(launch_blockers_details)
        issues_total_count_list.append(issues_total_count_list)
        issues_medium_count_list.append(issues_medium_count)
        issues_low_count_list.append(issues_low_count_list)

        # Test Planning Information
        requirement_analysis_status_list.append(requirement_analysis_status)
        requirement_analysis_eta_list.append(requirement_analysis_eta)
        use_case_created_count_list.append(use_case_created_count)
        test_case_created_count_list.append(test_case_created_count)
        use_case_created_status_list.append(use_case_created_status)
        test_case_created_status_list.append(test_case_created_status)
        use_case_created_eta_list.append(use_case_created_eta)
        test_case_created_eta_list.append(test_case_created_eta)

        # General Information
        test_data_generation_count_list.append(test_data_generation_count)
        adhoc_to_tc_conversion_count_list.append(adhoc_to_tc_conversion_count)
        adhoc_testing_bugs_found_count_list.append(adhoc_testing_bugs_found_count)
        bugs_verified_count_list.append(bugs_verified_count)
        test_case_optimized_count_list.append(test_case_optimized_count)
        test_data_generation_status_list.append(test_data_generation_status)
        adhoc_to_tc_conversion_status_list.append(adhoc_to_tc_conversion_status)
        adhoc_testing_bugs_found_status_list.append(adhoc_testing_bugs_found_status)
        bugs_verified_status_list.append(bugs_verified_status)
        test_case_optimized_status_list.append(test_case_optimized_status)
        test_data_generation_eta_list.append(test_data_generation_eta)
        adhoc_to_tc_conversion_eta_list.append(adhoc_to_tc_conversion_eta)
        adhoc_testing_bugs_found_eta_list.append(adhoc_testing_bugs_found_eta)
        bugs_verified_eta_list.append(bugs_verified_eta)
        test_case_optimized_eta_list.append(test_case_optimized_eta)


with st.container():
    if st.button('Submit KTR Update'):
        submit_data()
        st.write("KTR Update Submitted Successfully!")

with st.expander("Updates Provided So Far (Reflects only the active instance data that's submitted!)"):
    if "Automation" in project:
        st.write('Summary')
        summary = pd.DataFrame(list(zip(project_poc_list, project_name_list, project_description_list,
                                        current_testing_phase_list, current_testing_phase_eta_list)),
                               columns=['Project POC', 'Project Name', 'Project Description',
                                        'Current Testing Phase', 'Current Testing Phase ETA / Comments'])
        df_0 = summary
        st.dataframe(df_0)

        st.write('Automation')
        automation = pd.DataFrame(list(
            zip(project_name_list, test_case_feasibility_count_list, test_case_scripting_count_list,
                test_scripts_cr_count_list,
                pipeline_integration_type_list, test_case_feasibility_status_list, test_case_scripting_status_list,
                test_scripts_cr_status_list, pipeline_integration_status_list, test_case_feasibility_eta_list,
                test_case_scripting_eta_list, test_scripts_cr_eta_list, pipeline_integration_eta_list)),
            columns=['Project Name', 'Test Case Feasibility Count', 'Test Case Scripting Count',
                     'Test Scripts CR Count', 'Pipeline Integration Type', 'Test Case Feasibility Status',
                     'Test Case Scripting Status', 'Test Scripts CR Status', 'Pipeline Integration Status',
                     'Test Case Feasibility ETA', 'Test Case Scripting ETA', 'Test Scripts CR ETA',
                     'Pipeline Integration ETA'])
        df_1 = automation
        st.dataframe(df_1)
    else:
        st.write('Summary')
        summary = pd.DataFrame(list(zip(project_poc_list, project_name_list, project_description_list,
                                        current_testing_phase_list, current_testing_phase_eta_list)),
                               columns=['Project POC', 'Project Name', 'Project Description',
                                        'Current Testing Phase', 'Current Testing Phase ETA / Comments'])
        df_2 = summary
        st.dataframe(df_2)

        st.write('Planning')
        planning = pd.DataFrame(list(zip(project_name_list, requirement_analysis_status_list,
                                         requirement_analysis_eta_list, use_case_created_count_list,
                                         test_case_created_count_list, use_case_created_status_list,
                                         test_case_created_status_list, use_case_created_eta_list,
                                         test_case_created_eta_list)), columns=['Project Name',
                                                                                'Requirement Analysis Status',
                                                                                'Requirement Analysis ETA',
                                                                                'Use Case Created Count',
                                                                                'Test Case Created Count',
                                                                                'Use Case Created Status',
                                                                                'Test Case Created Status',
                                                                                'Use Case Created ETA',
                                                                                'Test Case Created ETA'])
        df_3 = planning
        st.dataframe(df_3)

        st.write("General")
        general = pd.DataFrame(list(zip(project_name_list, test_data_generation_count_list,
                                        adhoc_to_tc_conversion_count_list, adhoc_testing_bugs_found_count_list,
                                        bugs_verified_count_list, test_case_optimized_count_list,
                                        test_data_generation_status_list,
                                        adhoc_to_tc_conversion_status_list, adhoc_testing_bugs_found_status_list,
                                        bugs_verified_status_list,
                                        test_case_optimized_status_list, test_data_generation_eta_list,
                                        adhoc_to_tc_conversion_eta_list,
                                        adhoc_testing_bugs_found_eta_list, bugs_verified_eta_list,
                                        test_case_optimized_eta_list)),
                               columns=['Project Name', 'Test Data Generation Count', 'Ad-Hoc to TC Conversion Count',
                                        'Ad-Hoc Testing Bugs Found Count', 'Bugs Verified Count',
                                        'Test Case Optimized Count',
                                        'Test Data Generation Status', 'Ad-Hoc to TC Conversion Status',
                                        'Ad-Hoc Testing Bugs Found Status',
                                        'Bugs Verified Status', 'Test Case Optimized Status',
                                        'Test Data Generation ETA',
                                        'Ad-Hoc to TC Conversion ETA', 'Ad-Hoc Testing Bugs Found ETA',
                                        'Bugs Verified ETA',
                                        'Test Case Optimized ETA'])
        df_4 = general
        st.dataframe(df_4)

        st.write("Issue")
        issue = pd.DataFrame(list(zip(project_name_list, issues_high_count_list,
                                      launch_blockers_details_list, issues_total_count_list, issues_medium_count_list,
                                      issues_low_count_list)), columns=['Project Name', 'Issue High Count',
                                                                        'Launch Blocker Details', 'Issue Total Count',
                                                                        'Issue Medium Count', 'Issue Low Count'])
        df_5 = issue
        st.dataframe(df_5)

        st.write("Additional")
        additional = pd.DataFrame(list(zip(project_name_list, slow_spots_list, low_lights_list,
                                           one_or_two_way_decision_list, important_callouts_list)),
                                  columns=['Project Name', 'Slow Spots', 'Low Lights', 'One/Two Way Door Decision',
                                           'Important Callouts'])
        df_6 = additional
        st.dataframe(df_6)

st.markdown("""---""")
st.write("Please feel free to reach out [sureraja@amazon.com](https://phonetool.amazon.com/users/sureraja) "
         "in case of any missing fields (* such as Project POC, Project Name, etc.)")
